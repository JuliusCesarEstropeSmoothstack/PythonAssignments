import re
import logging
from openpyxl.utils import get_column_letter


# Function to find a value in a column with possible defined starting row and ending row for search using regex
def find_in_column(sheet, column, regex, exact_case=False, start=0, end=1048576):  # end defined as max number of Excel rows
    logging.info(f'Searching for pattern {regex} in column {column}')

    current = ''
    row = start
    # loop until target is found or exceeded looking at desired columns or to end of worksheet
    while row <= end if end < sheet.max_row else sheet.max_row:
        current = str(sheet[f'{column}{row}'].value)

        if not exact_case:  # convert to lower to make case-insensitive if looking for not exact match
            current = current.lower()

        # Check that the regex matches cell content
        if re.search(regex, current):
            logging.info(f'Found {current} matching {regex} on {column}{row} in {sheet.title}')
            return f'{column}{row}'
        row += 1

    logging.warning(f'Failed to find {current} matching {regex} in column {column} from sheet {sheet.title}')

    return None


def find_in_row(sheet, row, regex, exact_case=False, start=0, end=1048576):  # end defined as max number of Excel rows
    logging.info(f'Searching for pattern {regex} in row {row}')

    current = ''
    column_number = start
    # loop until target is found or exceeded looking at desired columns or to end of worksheet
    while column_number <= end if end < sheet.max_row else sheet.max_column:
        current = str(sheet[f'{get_column_letter(column_number)}{row}'].value)

        if not exact_case:  # convert to lower to make case-insensitive if looking for not exact match
            current = current.lower()

        # Check that the regex matches cell content
        if re.search(regex, current):
            logging.info(f'Found {current} matching {regex} on {get_column_letter(column_number)}{row}')
            return f'{get_column_letter(column_number)}{row}'
        column_number += 1

    logging.warning(f'Failed to find {current} matching {regex} in row {row} from sheet {sheet.title}')

    return None
