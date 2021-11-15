import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from month_enum import Month

logging.basicConfig(level=logging.INFO, filename='csvParserLog.log')

validInput = False
while not validInput:

    workbookName = input('Enter the file name')

    try:

        workbook = load_workbook(workbookName)

        logging.info(f'Attempting to parse {input} for a month and year in filename')
        month = workbookName.split('_')[3]
        year = workbookName.split('_')[4].replace('.xlsx', '')

        monthNumber = Month[month].value  # convert month string into number using Month enum

        logging.info(f'Successfully parsed inputname for month and year: {month}, {year}')
        validInput = True
    except FileExistsError:
        message = f'Workbook {workbookName} does not exist. Enter a different name.'
        logging.fatal(message)
        print(message)
    except IndexError:
        message = "Failed to parse a month and year from the inputname. '_' not properly used. Retry."
        logging.warning(message)
        print(message)
    except ValueError:
        message = f'{month} could not be parsed as a month. Please format as lowercase full month name.'
        logging.warning(message)
        print(message)

try:
    sheetName = 'Summary Rolling MoM'
    sheet = workbook[sheetName]  # grab sheet by name rather than the active sheet

    row = 2  # start from the second row, after the headers
    foundRow = None
    while row <= 13:  # loop until target is found or exceeded looking at row 13
        datetime = sheet[f'A{row}'].value  # value in column 'A' is in datetime format: YYyy-MM-DD HH:MM:SS
        date = str(datetime).split(' ')[0]  # split string into date and time parts, keeping date
        yearMonthDay = date.split('-')  # split year month and day to use later

        # Check that the month-year matches
        if int(yearMonthDay[1]) == monthNumber and yearMonthDay[0] == year:
            logging.info(f'Found {Month(monthNumber)} {year} on row {row}')
            foundRow = row
            break
        row += 1

    if foundRow is None:  # unable to find specified month-year within the first 100 rows of the sheet
        message = f'failed to find {Month(monthNumber)} {year} in column A of {workbookName}'
        logging.fatal(message)
        print(message)
        exit(-1)
    else:  # found the specified month-year, log and print out cell contents
        output = f"Calls offered: {sheet[f'B{row}'].value}"
        output += f"\n\tAbandon after 30s: {format(sheet[f'C{row}'].value * 100, '0.2f')}%"
        output += f"\n\tFCR: {format(sheet[f'D{row}'].value * 100, '0.2f')}%"
        output += f"\n\tDSAT: {format(sheet[f'E{row}'].value * 100, '0.2f')}%"
        output += f"\n\tCSAT: {format(sheet[f'F{row}'].value * 100, '0.2f')}%"
        logging.info(output)
        print(output)
except KeyError:
    message = f'{sheetName} not found in workbook {workbookName}. Skipping.'
    logging.warning(message)
    print(message)

try:
    sheetName = "VOC Rolling MoM"
    sheet = workbook[sheetName]
    column = 2
    foundRow = None
    while column <= 13:  # loop until target is found or exceeded looking at column 13
        try:
            datetime = sheet[f'{get_column_letter(column)}1'].value  # value in column 'A' is in datetime format: YYyy-MM-DD HH:MM:SS
            date = str(datetime).split(' ')[0]  # split string into date and time parts, keeping date
            yearMonthDay = date.split('-')  # split year month and day to use later

            # Check that the month-year matches
            if int(yearMonthDay[1]) == monthNumber and yearMonthDay[0] == year:
                logging.info(f'Found {Month(monthNumber)} {year} on row {row}')
                foundRow = row
                break
        except IndexError:
            # print(sheet[f'{get_column_letter(column)}1'].value)
            # print(month, str(sheet[f'{get_column_letter(column)}1'].value).title()[0:3])
            if month == str(sheet[f'{get_column_letter(column)}1'].value).lower():
                logging.info(f'Found {Month(monthNumber)} {year} on row {row}')
                foundRow = row

        column += 1

    if foundRow is None:  # unable to find specified month-year within the first 100 rows of the sheet
        message = f'failed to find {Month(monthNumber)} {year} in row 1 of {workbookName}'
        logging.fatal(message)
        print(message)
        exit(-1)
    else:  # found the specified month-year, log and print out cell contents
        # Promoters is on row 4
        output = 'In Net Promoter Score : '
        output += f"\n\tPromoters { '>= 200 : good' if sheet[f'{get_column_letter(column)}4'].value >= 200 else '< 200 : bad'}"
        # Passives is on row 6
        output += f"\n\tPassives { '>= 200 : good' if sheet[f'{get_column_letter(column)}6'].value >= 100 else '< 100 : bad'}"
        # Dectractors is on row 8
        output += f"\n\tDectractors { '>= 200 : good' if sheet[f'{get_column_letter(column)}8'].value >= 100 else '< 100 : bad'}"
        logging.info(output)
        print(output)
except KeyError:
    message = f'{sheetName} not found in workbook {workbookName}. Skipping.'
    logging.warning(message)
    print(message)
